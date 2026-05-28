#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# html2xlsx.py
#
# Convierte documentos HTML en archivos Excel con formato estructurado,
# soportando:
# - Encabezados y párrafos
# - Listas anidadas
# - Tablas complejas (rowspan/colspan)
# - Imágenes (base64, URL, local)
# - Enlaces internos y externos
#
# Author: Naidel
# Email: atmarquez@gmail.com
# Donate: https://paypal.me/atmarquez  # con PayPal
#
# Version: 1.0.0
# Copyright (C) 2026 Antonio Teodomiro Márquez Muñoz (Naidel)
#
# SPDX-License-Identifier: GPL-3.0-or-later
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# =============================================================================
# 1. METADATA
# =============================================================================

__author__ = "Antonio Teodomiro Márquez Muñoz (Naidel)"
__email__ = "atmarquez@gmail.com"
__version__ = "1.0.0"
__license__ = "GPL-3.0-or-later"
__donate__ = "https://paypal.me/atmarquez"

# =========================
# 2. IMPORTS
# =========================

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Side, Font
import requests
import os
import base64
import textwrap
import math
import logging
import argparse
import sys
import uuid

# =============================================================================
# 3. CONFIGURACIÓN GLOBAL
# =============================================================================
ANCHO_COLUMNA = 11  # valor por defecto
NUM_COLUMNAS = 10  # por defecto A-J
ANCHO_IMAGEN_MAX = None  # si no se define, se calcula automáticamente
ALTO_IMAGEN_MAX = None
IMAGE_FIT = "contain"
IMAGE_PADDING = 1
IMAGE_BACKGROUND = None
IMAGE_ALIGN = "left"
IMAGE_BORDER = False

wb = Workbook()
ws = wb.active

fila = 1
mapa_anclas = {}  # guarda id -> fila
enlaces_pendientes = []  # ✅ NUEVO: guardará los links internos

borde_fino = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

opciones = {
    "no_images": False,
    "no_tables": False,
    "no_links": False
}

# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.INFO,  # cambiar a DEBUG si quieres más detalle
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

# =========================
# 4. UTILIDADES GENERALES
# =========================

def ajustar_altura(texto, factor=1.6):
    """
    Calcula la altura de una fila de Excel en función del texto.

    La altura se estima en base al número de líneas que ocuparía
    el texto dentro del ancho disponible, aplicando un factor de
    ajuste para mejorar la visualización en Excel.

    Args:
        texto (str): Texto de la celda.
        factor (float, optional): Factor multiplicador para ajustar
            el espaciado vertical. Default es 1.6.

    Returns:
        float: Altura calculada para la fila.
    """
    if not texto:
        return 15

    # ✅ ancho real (ajustado para A-J)
    ancho_estimado = 180  # antes era ~100
    #160 → más compacto
    #180 → equilibrado ✅
    #200 → más espacioso

    lineas_reales = texto.split("\n")
    total = 0

    for linea in lineas_reales:
        if not linea.strip():
            total += 1
        else:
            total += max(1, len(textwrap.wrap(linea, ancho_estimado)))

    # ✅ margen extra para Excel (CLAVE)
    return total * 15 * factor

def altura_excel_real(texto, font_size=11, ancho_columnas=110, columnas=NUM_COLUMNAS):
    """
    Estima la altura real necesaria para visualizar correctamente
    un texto en Excel teniendo en cuenta el tamaño de fuente.

    Args:
        texto (str): Texto a medir.
        font_size (int, optional): Tamaño de fuente. Default es 11.
        ancho_columnas (int, optional): Ancho aproximado por columna.
        columnas (int, optional): Número de columnas usadas.

    Returns:
        float: Altura estimada en puntos.
    """
    if not texto:
        return 15

    # ✅ conversión aproximada Excel
    # 1 unidad de ancho ≈ 7 píxeles
    ancho_px = ancho_columnas * columnas * 7

    # ✅ ancho medio de carácter depende del tamaño
    # aprox: font_size * 0.6
    char_px = font_size * 0.6

    chars_por_linea = max(1, int(ancho_px / char_px))

    lineas = 0

    for linea in texto.split("\n"):
        if not linea.strip():
            lineas += 1
        else:
            lineas += max(1, math.ceil(len(linea) / chars_por_linea))

    # ✅ altura por línea (depende del tamaño fuente)
    altura_base = font_size * 1.5

    return lineas * altura_base

def insertar_espacio():
    global fila
    
    ws.row_dimensions[fila].height = 8  # altura pequeña tipo margen
    fila += 1

def archivo_en_uso(ruta):
    """
    Comprueba si un archivo está abierto o bloqueado.

    Args:
        ruta (str): Ruta del archivo.

    Returns:
        bool: True si está en uso, False en caso contrario.
    """
    if not os.path.exists(ruta):
        return False

    try:
        os.rename(ruta, ruta)
        return False
    except OSError:
        logger.warning("⚠️ El archivo '%s' está en uso", ruta)
        return True


# =========================
# 5. DETECCIÓN DE ESTILO
# =========================

def tiene_negrita(nodo):
    return nodo.find(["b", "strong"]) is not None

def tiene_cursiva(nodo):
    return nodo.find(["i", "em"]) is not None

def tiene_subrayado(nodo):
    return nodo.find("u") is not None

# =========================
# 6. ESCRITURA EN EXCEL
# =========================
    
def escribir_nodo(nodo, indent=0):
    """
    Escribe un nodo HTML (párrafo o encabezado) en la hoja Excel.

    Este método procesa el contenido textual del nodo, detecta
    enlaces y estilos, y lo inserta en una fila fusionada del Excel.
    También calcula la altura adecuada y aplica formato visual.

    Args:
        nodo (bs4.element.Tag): Nodo HTML a procesar.
        indent (int, optional): Nivel de indentación para listas
            o estructuras jerárquicas.

    Side Effects:
        - Modifica la hoja Excel (ws).
        - Incrementa la variable global 'fila'.
    """
    global fila
    global ANCHO_COLUMNA
    global NUM_COLUMNAS

    href = None
    link = nodo.find("a", href=True)

    texto = nodo.get_text("\n", strip=True) if nodo else ""

    if link:
        href = link.get("href")

    texto = " ".join(texto.split()) if texto else ""

    if not texto:
        return

    # ✅ Si hay múltiples enlaces, usar procesamiento avanzado
    if len(nodo.find_all("a", href=True)) > 1:
        procesar_texto_con_enlaces(nodo, indent)
        return

    if link:
        href = link.get("href")

    texto = " ".join(texto.split())
    if not texto:
        return

    # ✅ espacio BEFORE solo para títulos
    if nodo.name in ["h1", "h2", "h3"]:
        insertar_espacio()

    espacios = "   " * indent
    texto_final = espacios + texto

    # ✅ fusionar columnas A-J
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=NUM_COLUMNAS)

    celda = ws.cell(row=fila, column=1, value=texto_final)
    
    logger.debug("ℹ️ Insertando texto en fila %s", fila)
    
    if href:
        alineacion_horizontal = "left"   # ✅ para links
    else:
        alineacion_horizontal = "justify"  # ✅ normal

    celda.alignment = Alignment(
        wrap_text=True,
        vertical="top",
        horizontal=alineacion_horizontal
    )
    
    # ✅ aplicar enlace
    if href and not opciones["no_links"]:
        if href.startswith("http"):
            celda.hyperlink = href
            celda.font = Font(color="0000FF", underline="single")

        elif href.startswith("#"):
            enlaces_pendientes.append((fila, href))

    if nodo.name in ["code", "pre"]:
        celda.font = Font(name="Courier New", size=10)
        celda.fill = PatternFill(start_color="FFEEEEEE", fill_type="solid")

    if nodo.name == "h1":
        celda.font = Font(size=20, bold=True)
    elif nodo.name == "h2":
        celda.font = Font(size=16, bold=True)
    elif nodo.name == "h3":
        celda.font = Font(size=14, bold=True)
    elif nodo.name == "h4":
        celda.font = Font(size=13, bold=True)
    elif nodo.name == "h5":
        celda.font = Font(size=12, bold=True)
    elif nodo.name == "h6":
        celda.font = Font(size=11, bold=True)
    else:
        celda.font = Font(
            bold=tiene_negrita(nodo),
            italic=tiene_cursiva(nodo),
            underline="single" if tiene_subrayado(nodo) else None,
            size=11
        )       
        
    if nodo.name in ["h1", "h2", "h3", "h4"]:

        if nodo.name == "h1":
            estilo = "thick"
            top = Side(style="medium")
        elif nodo.name == "h2":
            estilo = "medium"
            top = None
        else:
            estilo = "thin"
            top = None

        for col in range(1, NUM_COLUMNAS + 1):
            ws.cell(row=fila, column=col).border = Border(
                top=top,
                bottom=Side(style=estilo)
            )

    # detectar tamaño
    if nodo.name == "h1":
        size = 20
    elif nodo.name == "h2":
        size = 16
    elif nodo.name == "h3":
        size = 14
    elif nodo.name == "h4":
        size = 13
    elif nodo.name == "h5":
        size = 12
    elif nodo.name == "h6":
        size = 11
    else:
        size = 11

    altura = altura_excel_real(
        texto_final,
        font_size=size,
        ancho_columnas=ANCHO_COLUMNA,  # ancho real por columna
        columnas=NUM_COLUMNAS
    )

    # pequeño margen adicional
    altura *= 1.15

    ws.row_dimensions[fila].height = altura
    fila += 1
    
    # ✅ añadir espacio SOLO después de cerrar la fila correcta
    if nodo.name in ["h1", "h2", "h3"]:        
        insertar_espacio()

def escribir_linea_simple(texto, indent=0, href=None):
    global fila
    global ANCHO_COLUMNA
    global NUM_COLUMNAS

    espacios = "   " * indent
    texto_final = espacios + texto

    # ✅ fusionar columnas A-J
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=NUM_COLUMNAS)

    celda = ws.cell(row=fila, column=1, value=texto_final)
    
    if href:
        alineacion_horizontal = "left"   # ✅ para links
    else:
        alineacion_horizontal = "justify"  # ✅ normal

    celda.alignment = Alignment(
        wrap_text=True,
        vertical="top",
        horizontal=alineacion_horizontal
    )

    # ✅ aplicar enlace
    if href and not opciones["no_links"]:
        if href.startswith("http"):
            celda.hyperlink = href
            celda.font = Font(color="0000FF", underline="single")

        elif href.startswith("#"):
            enlaces_pendientes.append((fila, href))
                
    altura = altura_excel_real(
        texto_final,
        font_size=11,
        ancho_columnas=ANCHO_COLUMNA
    )

    altura *= 1.1
    ws.row_dimensions[fila].height = altura
    fila += 1


# =========================
# 7. PROCESAMIENTO HTML
# =========================

def registrar_anclas(nodo):
    """
    Recorre el HTML para registrar todos los IDs presentes.

    Estos IDs se utilizan posteriormente para generar enlaces
    internos dentro del documento Excel.

    Args:
        nodo (bs4.element.Tag): Nodo HTML inicial.

    Side Effects:
        - Llena el diccionario global 'mapa_anclas'.
    """
    # ✅ solo procesar nodos HTML, no texto
    if not hasattr(nodo, "children"):
        return

    if nodo.has_attr("id"):
        mapa_anclas[nodo["id"]] = None  # destino provisional

    for hijo in nodo.children:
        registrar_anclas(hijo)
        
def procesar_lista(lista, nivel=0, numerada=False):
    """
    Procesa una lista HTML (<ul> o <ol>) y la convierte en filas de Excel.

    Gestiona correctamente listas anidadas, aplicando indentación y
    formatos de numeración o viñetas.

    Args:
        lista (bs4.element.Tag): Nodo de lista.
        nivel (int, optional): Nivel de indentación.
        numerada (bool, optional): Indica si es una lista ordenada.

    Side Effects:
        - Inserta múltiples filas en Excel.
    """
    global fila
    
    contador = 1

    for li in lista.find_all("li", recursive=False):

        # ✅ FIX: EXTRAER SOLO TEXTO DEL NIVEL ACTUAL (sin sublistas)
        partes = []

        for hijo in li.contents:
            if hasattr(hijo, "name") and hijo.name in ["ul", "ol"]:
                continue
            partes.append(str(hijo))

        href = None
        link = li.find("a", href=True)

        texto = BeautifulSoup("".join(partes), "html.parser").get_text(" ", strip=True)
        href = None

        link = li.find("a", href=True)
        if link:
            href = link.get("href")

        # limpiar espacios raros
        texto = texto.replace("\xa0", " ").strip()
        texto = " ".join(texto.split())

        if not texto:
            continue

        prefijo = f"{contador}. " if numerada else "• "
        if numerada:
            contador += 1

        escribir_linea_simple(prefijo + texto, nivel, href)

        # procesar sublistas correctamente
        for sub in li.find_all(["ul", "ol"], recursive=False):
            procesar_lista(sub, nivel + 1, sub.name == "ol")

    insertar_espacio()

def procesar_tabla(tabla):
    """
    Convierte una tabla HTML en una estructura de tabla en Excel.

    Soporta celdas con rowspan y colspan, aplica formato visual
    y ajusta automáticamente la altura de las filas.

    Args:
        tabla (bs4.element.Tag): Nodo <table> del HTML.

    Side Effects:
        - Inserta múltiples celdas y filas en Excel.
    """
    global fila
    global ANCHO_COLUMNA
    global NUM_COLUMNAS

    fila_inicio = fila
    grid = {}

    row_idx = 0

    for tr in tabla.find_all("tr"):
        col_idx = 0
        
        texto_fila = []

        while (row_idx, col_idx) in grid:
            col_idx += 1

        for celda_html in tr.find_all(["td", "th"]):

            colspan = int(celda_html.get("colspan", 1))
            rowspan = int(celda_html.get("rowspan", 1))

            texto = celda_html.get_text(" ", strip=True)
            texto_fila.append(texto)

            celda_excel = ws.cell(
                row=fila_inicio + row_idx,
                column=col_idx + 1,
                value=texto
            )

            celda_excel.alignment = Alignment(wrap_text=True, vertical="top")
            celda_excel.border = borde_fino

            # ✅ FORMATO (como fuera de tabla)
            celda_excel.font = Font(
                bold=celda_html.find(["b", "strong"]) is not None or celda_html.name == "th",
                italic=celda_html.find(["i", "em"]) is not None,
                underline="single" if celda_html.find("u") else None,
                size=11
            )

            # ✅ COLOR DE FONDO (si existe)
            style = celda_html.get("style", "")
            if "background-color" in style:
                try:
                    color = style.split("background-color:")[1].split(";")[0].strip()
                    color = color.replace("#", "")
                    if len(color) == 6:
                        color = "FF" + color.upper()

                    celda_excel.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"
                    )
                except:
                    logger.debug("⚠️ No se pudo aplicar background-color '%s': %s", style, e)

            else:
                logger.debug("ℹ️ Celda sin color de fondo")

            for r in range(rowspan):
                for c in range(colspan):
                    grid[(row_idx + r, col_idx + c)] = True

            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=fila_inicio + row_idx,
                    start_column=col_idx + 1,
                    end_row=fila_inicio + row_idx + rowspan - 1,
                    end_column=col_idx + colspan
                )

            col_idx += colspan

        texto_total = " ".join(texto_fila)
        ws.row_dimensions[fila_inicio + row_idx].height = ajustar_altura(texto_total)
        row_idx += 1

    fila = fila_inicio + row_idx + 1

    insertar_espacio()
    for col in range(1, NUM_COLUMNAS + 1):
        letra = chr(64 + col)  # A=65
        ws.column_dimensions[letra].width = ANCHO_COLUMNA

def procesar_imagen(src):
    """
    Inserta una imagen en la hoja de Excel desde distintas fuentes.

    Soporta:
        - Imágenes base64
        - URLs HTTP
        - Archivos locales

    La imagen se redimensiona para ajustarse al ancho del documento.

    Args:
        src (str): Ruta o fuente de la imagen.

    Side Effects:
        - Inserta imagen en Excel.
        - Avanza varias filas según el tamaño.
    """
    global fila

    try:
        logger.debug("ℹ️ Insertando imagen desde: %s", src)
        
        if src.startswith("data:image"):
            header, encoded = src.split(",", 1)
            data = base64.b64decode(encoded)

            ruta_temp = f"temp_img_{fila}.png"
            with open(ruta_temp, "wb") as f:
                f.write(data)

            img = XLImage(ruta_temp)

        elif src.startswith("http"):
            response = requests.get(src)

            ruta_temp = f"temp_img_{fila}.png"
            with open(ruta_temp, "wb") as f:
                f.write(response.content)

            img = XLImage(ruta_temp)

        else:
            ruta = os.path.join(os.path.dirname("documento.html"), src)
            img = XLImage(ruta)

        # ✅ PROPORCIÓN
        # ✅ calcular ancho máximo de la imagen       
        if ANCHO_IMAGEN_MAX:
            max_width = ANCHO_IMAGEN_MAX
        else:
            # calcular ancho en función de columnas y ancho de columna
            max_width = NUM_COLUMNAS * ANCHO_COLUMNA * 7

        # ✅ altura máxima (puedes dejarla fija o hacerla también configurable)
        if ALTO_IMAGEN_MAX:
            max_height = ALTO_IMAGEN_MAX
        else:
            # ✅ calcular alto coherente con el ancho
            max_width_base = NUM_COLUMNAS * ANCHO_COLUMNA * 7

            # proporcional tipo documento (A4 aprox relación 1:1.4)
            max_height = max_width_base * 1.4

        # ✅ Proporción imagen
        # Antes: ratio = min(max_width / img.width, max_height / img.height)
        # =========================
        # ESCALADO DE IMAGEN
        # =========================

        if IMAGE_FIT == "contain":
            ratio = min(max_width / img.width, max_height / img.height)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)

        elif IMAGE_FIT == "cover":
            ratio = max(max_width / img.width, max_height / img.height)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)

        elif IMAGE_FIT == "stretch":
            img.width = int(max_width)
            img.height = int(max_height)

        else:
            # fallback seguro
            ratio = min(max_width / img.width, max_height / img.height)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
    
        logger.debug("ℹ️ max_width: %s, max_height: %s", max_width, max_height)
        
        # ✅ calcular filas SIEMPRE
        filas_img = int(img.height / 15) + 1

        # ✅ background
        if IMAGE_BACKGROUND:
            color = IMAGE_BACKGROUND.replace("#", "").upper()

            for r in range(fila, fila + filas_img):
                for c in range(1, NUM_COLUMNAS + 1):
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color="FF" + color,
                        end_color="FF" + color,
                        fill_type="solid"
                    )

        # ✅ border (independiente del background)
        if IMAGE_BORDER:
            for r in range(fila, fila + filas_img):
                for c in range(1, NUM_COLUMNAS + 1):

                    celda = ws.cell(row=r, column=c)

                    borde = Border(
                        left=Side(style="medium") if c == 1 else None,
                        right=Side(style="medium") if c == NUM_COLUMNAS else None,
                        top=Side(style="medium") if r == fila else None,
                        bottom=Side(style="medium") if r == fila + filas_img - 1 else None,
                    )

                    celda.border = borde

        # ✅ padding antes
        fila += IMAGE_PADDING

        # ancho total en px del documento
        ancho_total = NUM_COLUMNAS * ANCHO_COLUMNA * 7
        ancho_img = img.width

        if IMAGE_ALIGN == "center":
            desplazamiento_px = max(0, (ancho_total - ancho_img) / 2)

        elif IMAGE_ALIGN == "right":
            desplazamiento_px = max(0, ancho_total - ancho_img)

        else:  # left
            desplazamiento_px = 0

        # convertir a columnas
        columna_offset = int(desplazamiento_px / (ANCHO_COLUMNA * 7))
        columna = 1 + columna_offset

        # evitar overflow
        columna = min(columna, NUM_COLUMNAS)

        letra_columna = chr(64 + columna)

        ws.add_image(img, f"{letra_columna}{fila}")

        filas_ocupadas = int(img.height / 15)

        # ✅ padding después
        fila += filas_ocupadas + IMAGE_PADDING

    except Exception as e:
        logger.error(
            "❌ Error procesando imagen '%s' en fila %s: %s",
            src,
            fila,
            e
        )


# =========================
# 8. MOTOR PRINCIPAL HTML
# =========================

def dentro_de_lista(nodo):
    parent = nodo.parent
    while parent is not None:
        if parent.name in ["ul", "ol"]:
            return True
        parent = parent.parent
    return False

def procesar_nodo(nodo):
    """
    Recorre recursivamente el árbol HTML y procesa cada nodo.

    Dependiendo del tipo de nodo:
        - Texto → escribir_nodo
        - Listas → procesar_lista
        - Tablas → procesar_tabla
        - Imágenes → procesar_imagen

    Args:
        nodo (bs4.element.Tag): Nodo HTML actual.

    Side Effects:
        - Genera contenido en la hoja Excel.
    """
    global fila
    global NUM_COLUMNAS
    
    try:
        if nodo.name is None:
            return

        logger.debug("ℹ️ Procesando nodo: %s", nodo.name)

        # ✅ guardar ID como destino de enlaces internos
        if nodo.has_attr("id"):
            mapa_anclas[nodo["id"]] = fila

        if dentro_de_lista(nodo) and nodo.name not in ["ul", "ol"]:
            return

        if nodo.name in ["p", "h1", "h2", "h3", "h4", "h5", "h6"]:
            escribir_nodo(nodo)

        elif nodo.name == "ul":
            procesar_lista(nodo)
            return

        elif nodo.name == "ol":
            procesar_lista(nodo, numerada=True)
            return

        elif nodo.name == "blockquote":
            escribir_nodo(nodo, indent=1)

        elif nodo.name == "table":
            if opciones["no_tables"]:
                return

            procesar_tabla(nodo)
            return

        elif nodo.name == "img":
            if opciones["no_images"]:
                return

            src = nodo.get("src")
            if src:
                procesar_imagen(src)

        elif nodo.name == "hr":
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)

            for col in range(1, NUM_COLUMNAS + 1):
                ws.cell(row=fila, column=col).border = Border(
                    bottom=Side(style="medium")
                )

            ws.row_dimensions[fila].height = 10
            fila += 1
            insertar_espacio()


        for hijo in nodo.children:
            procesar_nodo(hijo)


    except Exception as e:
        logger.error(
            "❌ Error procesando nodo '%s': %s",
            nodo.name if hasattr(nodo, "name") else "unknown",
            e
        )

def procesar_texto_con_enlaces(nodo, indent=0):
    """
    Procesa un nodo HTML que puede contener múltiples enlaces dentro del texto.
    Divide el contenido en fragmentos (texto normal + enlaces) y los escribe
    en filas separadas manteniendo el orden.

    Args:
        nodo (bs4.element.Tag): Nodo HTML a procesar.
        indent (int): Nivel de indentación para listas u otros contextos.

    Returns:
        None
    """
    global fila

    # Recorremos los hijos directos del nodo (mezcla de texto y etiquetas)
    for elemento in nodo.contents:

        # ---- TEXTO PLANO ----
        if isinstance(elemento, str):
            texto = elemento.strip()

            if texto:
                escribir_linea_simple(texto, indent)

        # ---- ENLACES ----
        elif elemento.name == "a":
            texto = elemento.get_text(" ", strip=True)
            href = elemento.get("href")

            if texto:
                escribir_linea_simple(texto, indent, href)

        # ---- OTROS ELEMENTOS (bold, italic, etc.) ----
        else:
            texto = elemento.get_text(" ", strip=True)
            if texto:
                escribir_linea_simple(texto, indent)


# =========================
# 9. ARGUMENTOS)
# =========================

def parse_arguments():
    """
    Parsea los argumentos de línea de comandos.

    Permite especificar:
        - Archivo de entrada HTML
        - Archivo de salida Excel (opcional)
        - Mostrar información de versión
        - Mostrar información de la licencia GNU v3
        - Especificar el nombre de la hoja Excel de salida
        - Activa el modo debug (muestra información detallada)
        - Activa el modo quiet (muestra solo mensajes de error)
        - Activa el modo para no incluir imágenes en el Excel
        - Activa el modo para no incluir tablas en el Excel
        - Activa el modo para desactivar enlaces en el Excel
        - Especificar el ancho de las columnas de Excel (por defecto: 11)
        - Especificar el número de columnas a fusionar (1-26). Por defecto: 10
        - Especificar ancho máximo de las imágenes en píxeles (mantiene proporción)
        - Especificar altura máxima de las imágenes en píxeles (mantiene proporción)
        - Controla cómo se ajusta la imagen dentro del espacio disponible (contain, cover, stretch)
        - Configura el número de filas de margen arriba y abajo de la imagen
        - Configura el color de fondo para el área de la imagen (hex, ej: FFFFFF)
        - Alineación de imágenes (left, center, right)
        - Añade un borde alrededor de la imagen
        - Muestra la ayuda
        
    Returns:
        argparse.Namespace: Objeto que contiene los argumentos introducidos.
    """
    parser = argparse.ArgumentParser(
        prog="html2xlsx",
        description="Convierte HTML a Excel (.xlsx)"
    )

    parser.add_argument(
        "input",
        nargs="?",
        help="Archivo HTML de entrada (html, htm, xhtml, mhtml)"
    )

    parser.add_argument(
        "output",
        nargs="?",
        help="Archivo Excel de salida (.xlsx)"
    )

    parser.add_argument(
        "-V", "--version",
        action="store_true",
        help="Muestra información de la versión y termina"
    )

    parser.add_argument(
        "-L", "--license",
        action="store_true",
        help="Muestra la licencia GNU GPL v3 y termina"
    )

    parser.add_argument(
        "--sheet",
        help="Nombre de la hoja Excel de salida (por defecto: Sheet)"
    )

    parser.add_argument(
        "--debug",
        action="store_true",
        help="Activa el modo debug (muestra información detallada)"
    )

    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Muestra solo mensajes de error"
    )

    parser.add_argument(
        "--no-images",
        action="store_true",
        help="No incluir imágenes en el Excel"
    )

    parser.add_argument(
        "--no-tables",
        action="store_true",
        help="No incluir tablas en el Excel"
    )

    parser.add_argument(
        "--no-links",
        action="store_true",
        help="Desactiva enlaces en el Excel"
    )        
    
    parser.add_argument(
        "--column-width",
        type=float,
        default=11,
        help="Ancho de las columnas de Excel (por defecto: 11)"
    )

    parser.add_argument(
        "--columns",
        type=int,
        default=10,
        help="Número de columnas a fusionar (1-26). Por defecto: 10"
    )

    parser.add_argument(
        "--image-max-width",
        type=int,
        help="Ancho máximo de las imágenes en píxeles (mantiene proporción)"
    )
    
    parser.add_argument(
        "--image-max-height",
        type=int,
        help="Altura máxima de las imágenes en píxeles"
    )
    
    parser.add_argument(
        "--image-fit",
        choices=["contain", "cover", "stretch"],
        default="contain",
        help="Modo de ajuste de imagen (contain, cover, stretch)"
    )

    parser.add_argument(
        "--image-padding",
        type=int,
        default=1,
        help="Número de filas de margen arriba y abajo de la imagen"
    )

    parser.add_argument(
        "--image-background",
        help="Color de fondo para el área de la imagen (hex, ej: FFFFFF)"
    )

    parser.add_argument(
        "--image-align",
        choices=["left", "center", "right"],
        default="left",
        help="Alineación de imágenes (left, center, right)"
    )

    parser.add_argument(
        "--image-border",
        action="store_true",
        help="Añade un borde alrededor de la imagen"
    )

    parser.add_argument(
        "--help-full",
        action="store_true",
        help="Muestra ayuda extendida con ejemplos de uso"
    )

    return parser.parse_args()

def mostrar_ayuda_completa():
    print("\n📘 html2xlsx - Convertidor de HTML a Excel\n")

    print("Descripción:")
    print(" Convierte documentos HTML en archivos Excel (.xlsx) manteniendo la estructura.")
    print(" Soporta texto formateado, listas, tablas complejas, imágenes y enlaces.\n")

    print("Uso:")
    print(" python html2xlsx.py <input.html> [output.xlsx] [opciones]\n")

    print("Opciones principales:")
    print(" --columns N               Número de columnas (1-26)")
    print(" --column-width N          Ancho de columnas")
    print(" --no-images               No incluir imágenes")
    print(" --no-tables               No incluir tablas")
    print(" --no-links                Desactivar enlaces\n")

    print("Opciones de imagen:")
    print(" --image-max-width N       Ancho máximo de imagen (px)")
    print(" --image-max-height N      Alto máximo de imagen (px)")
    print(" --image-fit MODE          Ajuste: contain, cover, stretch")
    print(" --image-padding N         Espacio arriba/abajo (filas)")
    print(" --image-background HEX    Color de fondo (ej: FFFFFF)")
    print(" --image-align MODE        Alineación: left, center, right")
    print(" --image-border            Añade borde exterior\n")

    print("Ejemplos:\n")

    print(" 1. Conversión básica:")
    print("    python html2xlsx.py archivo.html\n")

    print(" 2. Ajustar columnas:")
    print("    python html2xlsx.py archivo.html --columns 15 --column-width 12\n")

    print(" 3. Imágenes centradas con fondo:")
    print("    python html2xlsx.py archivo.html --image-align center --image-background EEEEEE\n")

    print(" 4. Imágenes con tamaño controlado:")
    print("    python html2xlsx.py archivo.html --image-max-width 800 --image-max-height 400\n")

    print(" 5. Layout profesional:")
    print("    python html2xlsx.py archivo.html \\")
    print("        --image-align center \\")
    print("        --image-background F8F8F8 \\")
    print("        --image-border\n")

    print(" 6. Sin imágenes ni tablas:")
    print("    python html2xlsx.py archivo.html --no-images --no-tables\n")

    print("\nNotas:")
    print(" - El archivo de salida se genera automáticamente si no se especifica.")
    print(" - Las imágenes mantienen proporción salvo 'stretch'.")
    print(" - El formato Excel se adapta automáticamente al contenido.\n")

    print("Autor:", __author__)
    print("Licencia:", __license__)
    print("")

def configurar_logging(debug=False, quiet=False):
    """
    Configura el sistema de logging según las opciones CLI.

    Prioridad:
        - debug → nivel DEBUG
        - quiet → nivel ERROR
        - por defecto → nivel INFO

    Args:
        debug (bool): Activa modo detallado.
        quiet (bool): Muestra solo errores.

    Side Effects:
        - Configura el logger global.
    """
    if debug:
        nivel = logging.DEBUG
    elif quiet:
        nivel = logging.ERROR
    else:
        nivel = logging.INFO

    logging.basicConfig(
        level=nivel,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True  # ✅ 🔥 ESTA ES LA CLAVE
    )

def obtener_nombre_hoja(nombre):
    """
    Valida y ajusta el nombre de la hoja de Excel.

    Reglas:
        - Máximo 31 caracteres
        - Elimina caracteres no permitidos

    Args:
        nombre (str|None): Nombre proporcionado por el usuario.

    Returns:
        str: Nombre válido para la hoja.
    """
    if not nombre:
        return "Sheet"

    # caracteres no permitidos en Excel
    invalidos = ['\\', '/', '*', '?', ':', '[', ']']

    for c in invalidos:
        nombre = nombre.replace(c, "")

    # limitar longitud
    return nombre[:31]
    
def mostrar_licencia():
    """
    Muestra la información de la licencia GNU GPL v3.

    Incluye:
        - Nombre de la licencia
        - Resumen de uso
        - Enlace oficial

    Side Effects:
        - Imprime información en consola.
        - No lanza excepciones.
    """
    print("\nLicencia: GNU General Public License v3.0\n")
    print("Este programa es software libre: puedes redistribuirlo y/o modificarlo")
    print("bajo los términos de la GNU General Public License según la Free Software Foundation.")
    print("Puedes usar la versión 3 de la licencia o cualquier versión posterior.\n")
    print("Este programa se distribuye SIN NINGUNA GARANTÍA.\n")
    print("Texto completo de la licencia:")
    print("https://www.gnu.org/licenses/gpl-3.0.html\n")
    print(f"Autor: {__author__}")
    print(f"Licencia declarada: {__license__}")
    
def mostrar_version():
    """
    Muestra la información de versión del programa.

    Incluye:
        - Versión
        - Autor
        - Email
        - Licencia
        - Enlace de donación

    Side Effects:
        - Imprime información en consola.
        - Termina la ejecución del programa.
    """
    print(f"html2xlsx versión {__version__}")
    print(f"Autor: {__author__}")
    print(f"Email: {__email__}")
    print(f"Licencia: {__license__}")
    print(f"Donaciones: {__donate__}")

def validar_archivo_entrada(ruta):
    """
    Valida el archivo de entrada.

    Comprueba:
        - Que exista
        - Que tenga extensión válida

    Args:
        ruta (str): Ruta del archivo de entrada.

    Raises:
        SystemExit: Si el archivo no es válido.
    """
    if not os.path.exists(ruta):
        logger.error("El archivo de entrada no existe: %s", ruta)
        sys.exit(1)

    extensiones_validas = (".html", ".htm", ".xhtml", ".mhtml")

    if not ruta.lower().endswith(extensiones_validas):
        logger.error("Extensión no válida: %s", ruta)
        sys.exit(1)

def obtener_salida(input_file, output_file):
    """
    Determina el nombre del archivo de salida.

    Reglas:
        - Si no se especifica salida → usa nombre del input
        - Si no tiene extensión → añade .xlsx

    Args:
        input_file (str): Archivo de entrada.
        output_file (str|None): Archivo de salida.

    Returns:
        str: Ruta final del archivo de salida.
    """
    if not output_file:
        base = os.path.splitext(input_file)[0]
        return base + ".xlsx"

    if not output_file.lower().endswith(".xlsx"):
        return output_file + ".xlsx"

    return output_file
 
def validar_sobrescritura(ruta):
    """
    Evita sobrescribir un archivo existente sin aviso.

    Args:
        ruta (str): Archivo de salida.

    Raises:
        SystemExit: Si el usuario decide no sobrescribir.
    """
    if os.path.exists(ruta):
        logger.warning("⚠️ El archivo ya existe: %s", ruta)

        respuesta = input("ℹ️ ¿Sobrescribir? (s/n): ").lower()

        if respuesta != "s":
            logger.info("❌ Operación cancelada por el usuario")
            sys.exit(0)

 
# =========================
# 10. MAIN (ejecución)
# =========================

def main():
    """
    Punto de entrada principal del programa.

    Realiza:
        - Parseo de argumentos CLI
        - Validación de entrada
        - Procesamiento HTML
        - Generación de Excel
    """
    global ANCHO_COLUMNA
    global NUM_COLUMNAS
    global ANCHO_IMAGEN_MAX
    global ALTO_IMAGEN_MAX
    global IMAGE_FIT
    global IMAGE_PADDING
    global IMAGE_BACKGROUND
    global IMAGE_ALIGN
    global IMAGE_BORDER
  
    args = parse_arguments()
    
    # ✅ configurar logging primero
    configurar_logging(args.debug, args.quiet)
    
    if args.help_full:
        mostrar_ayuda_completa()
        sys.exit(0)
    
    opciones["no_images"] = args.no_images
    opciones["no_tables"] = args.no_tables
    opciones["no_links"] = args.no_links

    if args.columns < 1 or args.columns > 26:
        logger.error("❌ El número de columnas debe estar entre 1 y 26")
        sys.exit(1)

    NUM_COLUMNAS = args.columns
    ANCHO_COLUMNA = args.column_width
    IMAGE_FIT = args.image_fit
    IMAGE_PADDING = args.image_padding
    IMAGE_BACKGROUND = args.image_background
    IMAGE_ALIGN = args.image_align
    IMAGE_BORDER = args.image_border

    if args.image_max_width is not None and args.image_max_width <= 0:
        logger.error("❌ El ancho de imagen debe ser mayor que 0")
        sys.exit(1)

    if args.image_max_height is not None and args.image_max_height <= 0:
        logger.error("❌ El alto de imagen debe ser mayor que 0")
        sys.exit(1)        

    if args.image_max_width:
        ANCHO_IMAGEN_MAX = args.image_max_width

    if args.image_max_height:
        ALTO_IMAGEN_MAX = args.image_max_height

        
    # ✅ 1. mostrar versión o licencia y salir
    if args.version:
        mostrar_version()
        sys.exit(0)

    if args.license:
        mostrar_licencia()
        sys.exit(0)

    # ✅ 2. comprobar que hay input
    if not args.input:
        logger.error("❌ Debes especificar un archivo de entrada. Usa --help para más información.")
        sys.exit(1)

    input_file = args.input
    output_file = obtener_salida(input_file, args.output)

    nombre_hoja = obtener_nombre_hoja(args.sheet)

    # ✅ aplicar nombre de hoja
    ws.title = nombre_hoja

    if not input_file:
        logger.error("❌ Debes especificar un archivo de entrada")
        sys.exit(1)

    validar_archivo_entrada(input_file)
    validar_sobrescritura(output_file)

    logger.info("ℹ️ Cargando archivo: %s", input_file)

    with open(input_file, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    logger.info("ℹ️ Alineación de imagen: %s", IMAGE_ALIGN)
    logger.info("ℹ️ Ancho de columna: %s", ANCHO_COLUMNA)
    logger.info("ℹ️ Número de columnas: %s", NUM_COLUMNAS)
    
    if ANCHO_IMAGEN_MAX:
        logger.info("ℹ️ Ancho máximo de imagen: %s px", ANCHO_IMAGEN_MAX)

    if ALTO_IMAGEN_MAX:
        logger.info("ℹ️ Alto máximo de imagen: %s px", ALTO_IMAGEN_MAX)

    logger.info("ℹ️ Registrando anclas...")
    registrar_anclas(soup.body)

    logger.info("ℹ️ Procesando contenido HTML...")
    procesar_nodo(soup.body)

    if not opciones["no_links"]:
        logger.info("ℹ️ Aplicando enlaces internos...")
        for fila_link, href in enlaces_pendientes:
            destino = href[1:]

            if destino in mapa_anclas and mapa_anclas[destino]:
                fila_destino = mapa_anclas[destino]

                celda = ws.cell(row=fila_link, column=1)
                celda.hyperlink = f"#'{ws.title}'!A{fila_destino}"
                celda.font = Font(color="0000FF", underline="single")

    if archivo_en_uso(output_file):
        logger.error("❌ El archivo está abierto: %s", output_file)
        sys.exit(1)

    logger.info("ℹ️ Guardando archivo: %s", output_file)
    wb.save(output_file)

    logger.info("ℹ️ Limpiando archivos temporales...")
    for archivo in os.listdir():
        if archivo.startswith("temp_img_") and archivo.endswith(".png"):
            try:
                os.remove(archivo)
            except Exception as e:
                logger.warning("❌ No se pudo borrar %s: %s", archivo, e)

    logger.info("✅ Proceso completado correctamente")


if __name__ == "__main__":
    main()